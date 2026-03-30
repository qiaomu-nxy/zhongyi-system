import io
import os

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse, StreamingResponse

from app.crud import patient as patient_crud
from app.crud import visit as visit_crud
from app.dependencies import DbSession, get_current_doctor
from app.exceptions import NotFoundError

router = APIRouter(prefix="/api/v1/backup", tags=["数据备份"])


@router.get("/export-excel", summary="导出全部数据为 Excel [需JWT]")
def export_excel(db: DbSession, _=Depends(get_current_doctor)):
    import openpyxl
    wb = openpyxl.Workbook()

    # 患者表
    ws = wb.active
    ws.title = "患者列表"
    ws.append(["ID", "姓名", "手机号", "性别", "出生日期", "过敏史", "注册时间"])
    patients = patient_crud.get_patients(db, limit=10000)
    for p in patients:
        ws.append([p.id, p.name, p.phone, p.gender, str(p.birth_date or ""), p.allergy_history or "", str(p.created_at)])

    # 就诊表
    ws2 = wb.create_sheet("就诊记录")
    ws2.append(["ID", "患者ID", "就诊日期", "第N次", "状态", "创建时间"])
    from app.models.visit import Visit
    visits = db.query(Visit).order_by(Visit.visit_date.desc()).limit(10000).all()
    for v in visits:
        ws2.append([v.id, v.patient_id, str(v.visit_date), v.visit_number, v.status, str(v.created_at)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=zhongyi_export.xlsx"},
    )


@router.get("/download-db", summary="下载数据库备份 [需JWT]")
def download_db(_=Depends(get_current_doctor)):
    db_path = "zhongyi.db"
    if not os.path.exists(db_path):
        raise NotFoundError("数据库文件不存在")
    return FileResponse(
        db_path,
        media_type="application/octet-stream",
        filename="zhongyi_backup.db",
    )


@router.get("/patients/{patient_id}/export", summary="导出单个患者病历 [需JWT]")
def export_patient(patient_id: int, db: DbSession, _=Depends(get_current_doctor)):
    import json, openpyxl
    patient = patient_crud.get_patient(db, patient_id)
    if not patient:
        raise NotFoundError("患者不存在")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者信息"
    ws.append(["姓名", "手机号", "性别", "出生日期", "过敏史"])
    ws.append([patient.name, patient.phone, patient.gender or "", str(patient.birth_date or ""), patient.allergy_history or ""])

    ws2 = wb.create_sheet("就诊记录")
    ws2.append(["就诊日期", "第N次", "状态", "诊断", "处方"])
    from app.models.visit import Visit
    from app.models.medical_record import MedicalRecord
    visits = db.query(Visit).filter(Visit.patient_id == patient_id).order_by(Visit.visit_date).all()
    for v in visits:
        rec = db.query(MedicalRecord).filter(MedicalRecord.visit_id == v.id).first()
        ws2.append([str(v.visit_date), v.visit_number, v.status,
                    rec.diagnosis if rec else "", rec.prescription if rec else ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"patient_{patient_id}_{patient.name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )